"""
GradCAS Applicant PDF Downloader — Extended Version
=====================================================
Downloads ALL document types per applicant into per-applicant subfolders:

    gradcas_downloads/
        LastName_FirstName/
            application.pdf
            cv.pdf
            personal_statement.pdf
            lor_Ref1.pdf
            lor_Ref2.pdf
            lor_Ref3.pdf

Full click path per applicant:
  [List page] Search last name → click row ">"
  → [Profile] click "Applications" in sidebar → click application row ">"
  → [Application] click "ATTACHMENTS" tab
      → extract application.pdf  (existing iframe title pattern)
      → extract cv.pdf           (div#cvresume)
      → extract personal_statement.pdf  (div#personal_statement)
  → click "RECOMMENDATIONS" tab
      → iterate div.reference-step entries in #ef_21311 and #ef_21312
      → extract lor_<Name>.pdf for each

S3 URL extraction strategy:
  Rather than clicking the pdf.js download button, we extract the direct
  S3 URL from the iframe src attribute and download with requests.
  This is more robust and bypasses the pdf.js layer entirely.
  Note: signed S3 URLs expire after 24 hours — download immediately.

Usage:
  1. Install dependencies (one time):
         pip install playwright openpyxl requests
         playwright install chromium

  2. Edit CONFIG below.

  3. Run:
         python download_gradcas_extended.py

  4. When the browser opens, log in via Pitt SSO, navigate to the
     Contacts / Applicants / All list page, then press Enter in terminal.
"""

import asyncio
import logging
import re
import sys
from pathlib import Path
from urllib.parse import urlparse, parse_qs, unquote

import openpyxl
import requests
from playwright.async_api import async_playwright, TimeoutError as PWTimeout

########################################################
# CONFIG
########################################################
EXCEL_PATH     = "applicants_2026.xlsx"
FIRST_NAME_COL = "First Name"
LAST_NAME_COL  = "Last Name"
DOWNLOAD_DIR   = Path("gradcas_downloads")
GRADCAS_URL    = "https://upitt-gradcas.admissionsbyliaison.com/"

TIMEOUT_MS     = 20_000
LOG_LEVEL      = logging.INFO
LOG_FILE       = "playwright_debug.log"

# Toggle which document types to download — set False to skip
DOWNLOAD_APPLICATION = True
DOWNLOAD_CV          = True
DOWNLOAD_PS          = True   # personal statement
DOWNLOAD_LORS        = True   # letters of recommendation

########################################################
# LOGGING SETUP
########################################################

def setup_logging():
    log_file = Path(LOG_FILE)
    root_logger = logging.getLogger()
    root_logger.handlers = []
    file_handler = logging.FileHandler(log_file, mode='a')
    console_handler = logging.StreamHandler(sys.stdout)
    formatter = logging.Formatter('%(asctime)s [%(levelname)s] %(message)s')
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)
    file_handler.setLevel(logging.DEBUG)
    console_handler.setLevel(LOG_LEVEL)
    root_logger.setLevel(logging.DEBUG)
    root_logger.addHandler(file_handler)
    root_logger.addHandler(console_handler)
    logger = logging.getLogger(__name__)
    logger.info(f"Logging initialized: Console={logging.getLevelName(LOG_LEVEL)}, File={log_file}")
    return logger

logger = setup_logging()

########################################################
# UTILITY FUNCTIONS
########################################################

def load_applicants(excel_path, first_col, last_col):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        fi = headers.index(first_col)
        li = headers.index(last_col)
    except ValueError as e:
        sys.exit(f"Column not found in Excel: {e}\nFound columns: {headers}")
    applicants = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        first = str(row[fi]).strip() if row[fi] else ""
        last  = str(row[li]).strip() if row[li] else ""
        if first or last:
            applicants.append({"first": first, "last": last})
    print(f"Loaded {len(applicants)} applicants from {excel_path}")
    return applicants


def safe_name(s):
    """Convert a string to a safe filename component."""
    return re.sub(r'[^\w\-]', '_', s.strip())


def applicant_dir(base_dir, first, last):
    """Return per-applicant subfolder, creating it if needed."""
    folder = base_dir / f"{safe_name(last)}_{safe_name(first)}"
    folder.mkdir(parents=True, exist_ok=True)
    return folder


async def wait_and_settle(page, ms=1500):
    logger.debug(f"  ⏳ waiting {ms}ms...")
    await page.wait_for_timeout(ms)


def extract_s3_url_from_src(src: str):
    """
    Extract the direct S3 URL from a pdf.js viewer src attribute.
    src format: /client/pdfjs/web/viewer.html?file=ENCODED_S3_URL#view=fitH
    Returns the decoded S3 URL string, or None on failure.
    """
    try:
        parsed = urlparse(src)
        file_param = parse_qs(parsed.query).get("file", [None])[0]
        return unquote(file_param) if file_param else None
    except Exception as e:
        logger.error(f"  ❌ Could not parse S3 URL: {e}")
        return None


def download_s3_to_file(s3_url: str, dest: Path) -> bool:
    """Download a signed S3 URL directly to dest using requests."""
    try:
        response = requests.get(s3_url, timeout=60)
        response.raise_for_status()
        dest.write_bytes(response.content)
        size_kb = dest.stat().st_size // 1024
        logger.info(f"  ✅ Saved: {dest.name} ({size_kb} KB)")
        print(f"      ✅ {dest.name}  ({size_kb} KB)")
        return True
    except Exception as e:
        logger.error(f"  ❌ S3 download failed for {dest.name}: {e}")
        print(f"      ❌ {dest.name} — download failed: {e}")
        return False


async def expand_and_get_s3_url(page, section_locator):
    """
    Ensure a section's pdf.js iframe is visible, then extract and return
    the S3 URL from it. Clicks the step-card-name toggle if needed.
    Returns (s3_url: str | None, already_existed: bool)
    """
    iframe = section_locator.locator("iframe[data-id='PDFViewer']")
    iframe_count = await iframe.count()

    if iframe_count == 0:
        # Section is collapsed — click the caret toggle to expand
        toggle = section_locator.locator("a.step-card-name")
        logger.debug("  Expanding collapsed section...")
        await toggle.click(timeout=TIMEOUT_MS)
        await wait_and_settle(page, ms=2000)
        iframe_count = await iframe.count()

    if iframe_count == 0:
        logger.warning("  ⚠️  No iframe found after expansion attempt")
        return None, False

    src = await iframe.first.get_attribute("src")
    if not src:
        logger.warning("  ⚠️  iframe src attribute is empty")
        return None, False

    return extract_s3_url_from_src(src), True


########################################################
# NAVIGATION FUNCTIONS (unchanged from original)
########################################################

async def go_back_to_list(page):
    logger.info("🔙 go_back_to_list")
    try:
        await page.locator("i[data-name='people']").click(timeout=TIMEOUT_MS)
        await wait_and_settle(page, ms=2000)
        try:
            clicked = await page.evaluate("""
                () => {
                    const icons = Array.from(document.querySelectorAll('i.material-icons'));
                    for (let icon of icons) {
                        if (icon.textContent.trim() === 'cancel') {
                            const style = window.getComputedStyle(icon);
                            if (style.display !== 'none' && style.visibility !== 'hidden') {
                                icon.click();
                                return true;
                            }
                        }
                    }
                    return false;
                }
            """)
            if clicked:
                await wait_and_settle(page, ms=1000)
        except Exception:
            pass
    except PWTimeout:
        logger.error("  ❌ Timeout: Could not find people/contacts sidebar icon")


async def search_and_open_applicant(page, first, last):
    logger.info(f"🔍 Searching for '{first} {last}'")
    try:
        await page.locator("i.search-button").wait_for(state="visible", timeout=30_000)
        await page.locator("i.search-button").click(timeout=TIMEOUT_MS)
        await page.wait_for_timeout(800)
        search_input = page.locator("input[type='text'], input[type='search']").last
        await search_input.wait_for(state="visible", timeout=TIMEOUT_MS)
        await search_input.click(timeout=TIMEOUT_MS)
        await search_input.fill(last)
        await search_input.press("Enter")
        await page.wait_for_timeout(3000)
    except PWTimeout as e:
        logger.error(f"  ❌ Timeout searching: {e}")
        return False
    except Exception as e:
        logger.error(f"  ❌ Error searching: {e}")
        return False

    rows = page.locator("tbody tr")
    count = await rows.count()
    if count == 0:
        logger.warning(f"  ⚠️  No results for '{last}'")
        print(f"  No results for last name '{last}'")
        return False

    target_row = None
    for i in range(count):
        row = rows.nth(i)
        text = (await row.text_content() or "").strip()
        if first.lower() in text.lower() and last.lower() in text.lower():
            target_row = row
            break

    if target_row is None:
        if count == 1:
            target_row = rows.first
        else:
            logger.warning(f"  ⚠️  '{first} {last}' not uniquely identified among {count} rows")
            return False

    try:
        await target_row.locator("td").last.click(timeout=8000)
    except PWTimeout:
        await target_row.click(timeout=8000)

    await wait_and_settle(page)
    return True


async def click_applications_sidebar(page):
    logger.info("📋 click_applications_sidebar")
    try:
        await page.locator("a[data-id='tab-applications']").click(timeout=TIMEOUT_MS)
        await wait_and_settle(page, ms=2000)
        return True
    except PWTimeout:
        logger.error("  ❌ Could not find Applications sidebar link")
        return False


async def click_application_row(page):
    logger.info("📄 click_application_row")
    try:
        await page.locator("td a, tbody tr a").first.click(timeout=TIMEOUT_MS)
        await wait_and_settle(page)
        return True
    except PWTimeout:
        try:
            await page.locator("tbody tr").first.locator("td").last.click(timeout=8000)
            await wait_and_settle(page)
            return True
        except PWTimeout:
            logger.error("  ❌ Could not click into application row")
            return False


async def click_attachments_tab(page):
    logger.info("📎 click_attachments_tab")
    try:
        await page.locator("text=ATTACHMENTS").click(timeout=TIMEOUT_MS)
        await wait_and_settle(page, ms=2000)
        return True
    except PWTimeout:
        logger.error("  ❌ Could not find ATTACHMENTS tab")
        return False


async def click_recommendations_tab(page):
    logger.info("📝 click_recommendations_tab")
    try:
        await page.locator("text=RECOMMENDATIONS").click(timeout=TIMEOUT_MS)
        await wait_and_settle(page, ms=2000)
        return True
    except PWTimeout:
        logger.error("  ❌ Could not find RECOMMENDATIONS tab")
        return False


########################################################
# DOCUMENT DOWNLOAD FUNCTIONS
########################################################

async def download_application_pdf(page, dest_dir):
    """Download the main application PDF (original logic, refactored to use S3 extraction)."""
    dest = dest_dir / "application.pdf"
    if dest.exists() and dest.stat().st_size > 0:
        print(f"      ⏭️  application.pdf already exists, skipping")
        return True

    logger.info("📄 download_application_pdf")
    try:
        iframe_locator = page.locator("iframe[data-id='PDFViewer'][title$='_application.pdf']")
        if await iframe_locator.count() == 0:
            await page.locator("a[data-id^='toggle-step']").first.click(timeout=TIMEOUT_MS)
            await wait_and_settle(page, ms=1000)

        await iframe_locator.wait_for(state="attached", timeout=TIMEOUT_MS)
        await wait_and_settle(page, ms=2000)

        src = await iframe_locator.first.get_attribute("src")
        s3_url = extract_s3_url_from_src(src)
        if not s3_url:
            return False
        return download_s3_to_file(s3_url, dest)

    except PWTimeout as e:
        logger.error(f"  ❌ Timeout in download_application_pdf: {e}")
        return False
    except Exception as e:
        logger.error(f"  ❌ Error in download_application_pdf: {e}")
        return False


async def download_cv(page, dest_dir):
    """Download CV/Resume from div#cvresume."""
    dest = dest_dir / "cv.pdf"
    if dest.exists() and dest.stat().st_size > 0:
        print(f"      ⏭️  cv.pdf already exists, skipping")
        return True

    logger.info("📄 download_cv")
    try:
        section = page.locator("div#cvresume")
        s3_url, found = await expand_and_get_s3_url(page, section)
        if not s3_url:
            print(f"      ⚠️  cv.pdf — could not extract URL")
            return False
        return download_s3_to_file(s3_url, dest)
    except Exception as e:
        logger.error(f"  ❌ Error in download_cv: {e}")
        return False


async def download_personal_statement(page, dest_dir):
    """Download Personal Statement from div#personal_statement."""
    dest = dest_dir / "personal_statement.pdf"
    if dest.exists() and dest.stat().st_size > 0:
        print(f"      ⏭️  personal_statement.pdf already exists, skipping")
        return True

    logger.info("📄 download_personal_statement")
    try:
        section = page.locator("div#personal_statement")
        s3_url, found = await expand_and_get_s3_url(page, section)
        if not s3_url:
            print(f"      ⚠️  personal_statement.pdf — could not extract URL")
            return False
        return download_s3_to_file(s3_url, dest)
    except Exception as e:
        logger.error(f"  ❌ Error in download_personal_statement: {e}")
        return False


async def download_lors(page, dest_dir):
    """
    Download all Letters of Recommendation.
    Iterates div.reference-step entries inside #ef_21311 and #ef_21312.
    Saves as lor_FirstName_LastName.pdf
    Returns list of booleans (one per LoR attempt).
    """
    logger.info("📝 download_lors")
    results = []

    for section_id in ["#ef_21311", "#ef_21312"]:
        section = page.locator(section_id)
        if await section.count() == 0:
            logger.debug(f"  Section {section_id} not present, skipping")
            continue

        recommenders = section.locator("div.reference-step")
        rec_count = await recommenders.count()
        logger.info(f"  Found {rec_count} recommender(s) in {section_id}")

        for i in range(rec_count):
            rec = recommenders.nth(i)

            # Extract recommender name for filename
            try:
                name_text = (await rec.locator("div.reference-name").first.text_content()).strip()
            except Exception:
                name_text = f"recommender_{i+1}"

            filename = f"lor_{safe_name(name_text)}.pdf"
            dest = dest_dir / filename

            if dest.exists() and dest.stat().st_size > 0:
                print(f"      ⏭️  {filename} already exists, skipping")
                results.append(True)
                continue

            print(f"      📥 {name_text}...")

            try:
                iframe = rec.locator("iframe[data-id='PDFViewer']")

                # Expand if not already open
                if await iframe.count() == 0:
                    await rec.locator("div.reference-chevron").click(timeout=TIMEOUT_MS)
                    await wait_and_settle(page, ms=2000)

                # Fallback: try clicking reference-card link
                if await iframe.count() == 0:
                    try:
                        await rec.locator("a.reference-card").click(timeout=5000)
                        await wait_and_settle(page, ms=2000)
                    except Exception:
                        pass

                if await iframe.count() == 0:
                    logger.warning(f"  ⚠️  No iframe for {name_text}")
                    print(f"      ⚠️  {filename} — no PDF viewer found")
                    results.append(False)
                    continue

                src = await iframe.first.get_attribute("src")
                s3_url = extract_s3_url_from_src(src)
                if not s3_url:
                    results.append(False)
                    continue

                success = download_s3_to_file(s3_url, dest)
                results.append(success)

                # Collapse before moving to next (keeps page clean)
                try:
                    await rec.locator("div.reference-chevron").click(timeout=5000)
                    await wait_and_settle(page, ms=500)
                except Exception:
                    pass

            except PWTimeout as e:
                logger.error(f"  ❌ Timeout for LoR {name_text}: {e}")
                print(f"      ❌ {filename} — timeout")
                results.append(False)
            except Exception as e:
                logger.error(f"  ❌ Error for LoR {name_text}: {e}")
                print(f"      ❌ {filename} — error: {e}")
                results.append(False)

    return results


########################################################
# MAIN
########################################################

async def main():
    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    applicants = load_applicants(EXCEL_PATH, FIRST_NAME_COL, LAST_NAME_COL)

    succeeded = []
    partial   = []
    failed    = []

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=False)
        context = await browser.new_context(
            accept_downloads=True,
            viewport={"width": 1400, "height": 900},
        )
        page = await context.new_page()
        page.on("framenavigated", lambda e: logger.info(f"🌐 NAV: {e.url}"))
        page.on("console", lambda m: logger.debug(f"📝 CONSOLE: {m.text}"))

        try:
            await page.goto(GRADCAS_URL)
        except Exception as e:
            logger.error(f"❌ Error loading initial page: {e}")

        print("\n" + "="*60)
        print("Browser is open. Please:")
        print("  1. Log in via Pitt SSO")
        print("  2. Navigate to: Contacts > Applicants > All")
        print("  3. Return here and press Enter")
        print("="*60 + "\n")

        try:
            input("Press Enter when ready > ")
        except (EOFError, KeyboardInterrupt):
            print("\n⚠️  Interrupted.")
            return

        for i, applicant in enumerate(applicants, 1):
            first   = applicant["first"]
            last    = applicant["last"]
            dest_dir = applicant_dir(DOWNLOAD_DIR, first, last)

            print(f"\n[{i}/{len(applicants)}] {first} {last}")
            print(f"  📁 {dest_dir.name}/")

            doc_results = {}

            try:
                logger.info("="*60)
                logger.info(f"Processing: {first} {last}")

                await go_back_to_list(page)

                if not await search_and_open_applicant(page, first, last):
                    failed.append(f"{first} {last} — not found"); continue

                if not await click_applications_sidebar(page):
                    failed.append(f"{first} {last} — Applications sidebar missing"); continue

                if not await click_application_row(page):
                    failed.append(f"{first} {last} — could not open application"); continue

                # ── ATTACHMENTS TAB ──────────────────────────────────────
                if not await click_attachments_tab(page):
                    failed.append(f"{first} {last} — ATTACHMENTS tab missing"); continue

                if DOWNLOAD_APPLICATION:
                    print("    Application PDF...")
                    doc_results["application"] = await download_application_pdf(page, dest_dir)

                if DOWNLOAD_CV:
                    print("    CV/Resume...")
                    doc_results["cv"] = await download_cv(page, dest_dir)

                if DOWNLOAD_PS:
                    print("    Personal Statement...")
                    doc_results["personal_statement"] = await download_personal_statement(page, dest_dir)

                # ── RECOMMENDATIONS TAB ──────────────────────────────────
                if DOWNLOAD_LORS:
                    print("    Letters of Recommendation...")
                    if await click_recommendations_tab(page):
                        lor_results = await download_lors(page, dest_dir)
                        doc_results["lors_ok"] = all(lor_results) if lor_results else False
                    else:
                        doc_results["lors_ok"] = False

                # ── RESULT ───────────────────────────────────────────────
                all_ok = all(v for v in doc_results.values() if isinstance(v, bool))
                if all_ok:
                    print(f"  ✅ Complete")
                    succeeded.append(f"{first} {last}")
                else:
                    issues = [k for k, v in doc_results.items() if v is False]
                    print(f"  ⚠️  Partial — issues with: {issues}")
                    partial.append(f"{first} {last} — {issues}")

            except Exception as e:
                print(f"  ❌ Unexpected error: {e}")
                logger.exception(f"Unexpected error for {first} {last}")
                failed.append(f"{first} {last} — error: {e}")

        await browser.close()

    print("\n" + "="*60)
    print(f"COMPLETE:")
    print(f"  ✅ Fully succeeded : {len(succeeded)}")
    print(f"  ⚠️  Partial        : {len(partial)}")
    print(f"  ❌ Failed          : {len(failed)}")
    if partial:
        print("\nPartial (re-run to retry — already downloaded files will be skipped):")
        for n in partial:
            print(f"  - {n}")
    if failed:
        print("\nFailed:")
        for n in failed:
            print(f"  - {n}")
    print(f"\nFiles saved to: {DOWNLOAD_DIR.resolve()}")
    print("="*60)


if __name__ == "__main__":
    asyncio.run(main())